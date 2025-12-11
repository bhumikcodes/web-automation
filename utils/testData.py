import openpyxl


def test_data():
    workbook = openpyxl.load_workbook("TestData.xlsx")
    sheet = workbook["Username"]
    values = []
    for row in range(2,sheet.max_row+1):
        input_1 = sheet.cell(row=row,column=1).value
        print(input_1)
        values.append(input_1)
    return(values)

